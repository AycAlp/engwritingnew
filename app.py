import os
import io
import csv
import uuid
import zipfile
import statistics
from datetime import datetime
from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_from_directory, send_file, abort)
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
from mfrm import compute_mfrm
from database import (
    init_db, get_db, get_user_by_email, verify_password, UPLOAD_DIR, READING_DIR, TRAINING_DIR,
    get_all_essays, get_essay, get_essay_images,
    get_scores_for_user, get_all_scores, get_gold_standard,
    get_all_instructors, get_submission_count, get_score_for_user_essay,
    get_phase, set_phase,
    get_next_essay_for_rater, get_rater_position_info, get_rater_assignments,
    generate_assignments_for_rater, assign_new_essay_to_all_raters,
    recompute_instructor_metrics, generate_author_tag, generate_system_id,
    create_standardization, get_standardization, get_all_standardizations,
    get_active_standardizations, archive_standardization, reopen_standardization,
    get_essays_by_standardization, get_cross_standardization_stats,
    get_training_materials, add_training_material, delete_training_material,
    get_training_notes, add_training_note, update_training_note_position, delete_training_note,
)

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "eng101-writing-standardization-2025")
app.config["MAX_CONTENT_LENGTH"] = 64 * 1024 * 1024  # 64 MB

# Run DB migrations on every startup (safe to call repeatedly)
init_db()

ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "gif", "webp", "pdf"}
ALLOWED_READING_EXTENSIONS = {"pdf", "doc", "docx", "txt", "jpg", "jpeg", "png"}
ALLOWED_TRAINING_EXTENSIONS = {"mp4", "webm", "mov", "avi", "mp3", "wav", "m4a", "ogg", "pdf", "doc", "docx"}

CATEGORIES = ["task_req", "argument", "support", "language", "readability", "formatting"]
CATEGORY_LABELS = {
    "task_req":    "Task Requirements",
    "argument":    "Argument",
    "support":     "Support",
    "language":    "Use of Language",
    "readability": "Readability",
    "formatting":  "Formatting",
}
CATEGORY_SCALES = {
    "task_req":    {"min": 0, "max": 2, "labels": {0: "Does Not Meet", 1: "Partially Meets", 2: "Meets"}},
    "argument":    {"min": 1, "max": 5, "labels": {1: "Poor", 2: "Unsatisfactory", 3: "Satisfactory", 4: "Good", 5: "Excellent"}},
    "support":     {"min": 1, "max": 5, "labels": {1: "Poor", 2: "Unsatisfactory", 3: "Satisfactory", 4: "Good", 5: "Excellent"}},
    "language":    {"min": 1, "max": 5, "labels": {1: "Poor", 2: "Unsatisfactory", 3: "Satisfactory", 4: "Good", 5: "Excellent"}},
    "readability": {"min": 1, "max": 3, "labels": {1: "Poor", 2: "Unsatisfactory", 3: "Satisfactory"}},
    "formatting":  {"min": 0, "max": 2, "labels": {0: "Does Not Meet", 1: "Partially Meets", 2: "Meets"}},
}
MAX_TOTAL = 22

CERTAINTY_LABELS = {
    5: "Completely certain",
    4: "Mostly certain",
    3: "Somewhat certain",
    2: "Mostly uncertain",
    1: "Very uncertain",
}

PHASES = {
    "data_collection": "Data Collection",
    "analysis":        "Analysis",
    "reporting":       "Reporting",
}


# ── Utilities ─────────────────────────────────────────────────────────────────

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def allowed_reading(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_READING_EXTENSIONS


def allowed_training(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_TRAINING_EXTENSIONS


def training_file_type(filename):
    ext = filename.rsplit(".", 1)[1].lower() if "." in filename else ""
    if ext in {"mp4", "webm", "mov", "avi"}:
        return "video"
    if ext in {"mp3", "wav", "m4a", "ogg"}:
        return "audio"
    if ext == "pdf":
        return "pdf"
    if ext in {"doc", "docx"}:
        return "word"
    return "other"


def score_total(row):
    return sum(row[c] or 0 for c in CATEGORIES)


def compute_icc(ratings_matrix):
    try:
        import numpy as np
        data = np.array(ratings_matrix, dtype=float)
        n, k = data.shape
        if n < 2 or k < 2:
            return None
        grand_mean = data.mean()
        ss_total = ((data - grand_mean) ** 2).sum()
        row_means = data.mean(axis=1, keepdims=True)
        ss_rows = k * ((row_means - grand_mean) ** 2).sum()
        col_means = data.mean(axis=0, keepdims=True)
        ss_cols = n * ((col_means - grand_mean) ** 2).sum()
        ss_error = ss_total - ss_rows - ss_cols
        ms_rows = ss_rows / (n - 1)
        ms_error = ss_error / ((n - 1) * (k - 1))
        icc = (ms_rows - ms_error) / (ms_rows + (k - 1) * ms_error)
        return round(float(icc), 3)
    except Exception:
        return None


def compute_reliability_stats():
    all_scores = get_all_scores()
    essays = get_all_essays()
    instructors = get_all_instructors()
    if not all_scores or not instructors:
        return None

    matrix = {}
    for s in all_scores:
        matrix.setdefault(s["essay_id"], {})[s["user_id"]] = s

    instructor_ids = [i["id"] for i in instructors]
    essay_ids = [e["id"] for e in essays]

    # Per-instructor stats
    instructor_stats = {}
    for instr in instructors:
        uid = instr["id"]
        totals = [score_total(matrix[eid][uid])
                  for eid in essay_ids if eid in matrix and uid in matrix[eid]]
        if not totals:
            continue
        certainties = [matrix[eid][uid]["certainty"]
                       for eid in essay_ids
                       if eid in matrix and uid in matrix[eid]
                       and matrix[eid][uid]["certainty"] is not None]
        instructor_stats[uid] = {
            "name": instr["name"],
            "email": instr["email"],
            "bilkent_id": instr["bilkent_id"],
            "mean_total": round(statistics.mean(totals), 2),
            "essays_marked": len(totals),
            "totals": totals,
            "avg_certainty": round(statistics.mean(certainties), 2) if certainties else None,
            "cat_means": {},
        }

    # Category means per instructor
    for uid in instructor_ids:
        if uid not in instructor_stats:
            continue
        for cat in CATEGORIES:
            vals = [matrix[eid][uid][cat]
                    for eid in essay_ids
                    if eid in matrix and uid in matrix[eid]
                    and matrix[eid][uid][cat] is not None]
            instructor_stats[uid]["cat_means"][cat] = (
                round(statistics.mean(vals), 2) if vals else None
            )

    # Group means per category
    group_cat_means = {}
    for cat in CATEGORIES:
        all_cat_vals = [
            matrix[eid][uid][cat]
            for eid in essay_ids
            for uid in instructor_ids
            if eid in matrix and uid in matrix[eid] and matrix[eid][uid][cat] is not None
        ]
        group_cat_means[cat] = round(statistics.mean(all_cat_vals), 2) if all_cat_vals else None

    # Essay group means (for heatmap)
    essay_group_means = {}
    for eid in essay_ids:
        if eid in matrix:
            vals = [score_total(matrix[eid][uid]) for uid in matrix[eid]]
            if vals:
                essay_group_means[eid] = statistics.mean(vals)

    all_totals = [t for s in instructor_stats.values() for t in s["totals"]]
    overall_mean = round(statistics.mean(all_totals), 2) if all_totals else 0

    for uid, s in instructor_stats.items():
        s["leniency"] = round(s["mean_total"] - overall_mean, 2)
        if s["leniency"] > 1.0:
            s["tendency"] = "Lenient"; s["tendency_class"] = "lenient"
        elif s["leniency"] < -1.0:
            s["tendency"] = "Harsh"; s["tendency_class"] = "harsh"
        else:
            s["tendency"] = "Calibrated"; s["tendency_class"] = "calibrated"

        s["cat_leniency"] = {}
        for cat in CATEGORIES:
            my_mean = s["cat_means"].get(cat)
            grp_mean = group_cat_means.get(cat)
            if my_mean is not None and grp_mean is not None:
                s["cat_leniency"][cat] = round(my_mean - grp_mean, 2)
            else:
                s["cat_leniency"][cat] = None

    # ICC per category
    icc_per_category = {}
    for cat in CATEGORIES:
        ratings = []
        for eid in essay_ids:
            if eid not in matrix:
                continue
            row = [matrix[eid][uid][cat]
                   for uid in instructor_ids
                   if uid in matrix[eid] and matrix[eid][uid][cat] is not None]
            if len(row) == len(instructor_ids):
                ratings.append(row)
        icc_per_category[cat] = compute_icc(ratings) if ratings else None

    # Overall ICC
    total_ratings = []
    for eid in essay_ids:
        if eid not in matrix:
            continue
        row = [score_total(matrix[eid][uid])
               for uid in instructor_ids if uid in matrix[eid]]
        if len(row) == len(instructor_ids):
            total_ratings.append(row)
    overall_icc = compute_icc(total_ratings)

    # Heatmap rows
    heatmap = []
    for uid in instructor_ids:
        if uid not in instructor_stats:
            continue
        row = {"name": instructor_stats[uid]["name"], "cells": []}
        for eid in essay_ids:
            if eid in matrix and uid in matrix[eid]:
                t = score_total(matrix[eid][uid])
                dev = round(t - essay_group_means.get(eid, t), 1)
                row["cells"].append({"total": t, "deviation": dev, "has_score": True})
            else:
                row["cells"].append({"total": None, "deviation": None, "has_score": False})
        heatmap.append(row)

    # Standardized score comparison
    gs_comparison = {}
    for eid in essay_ids:
        gs = get_gold_standard(eid)
        if not gs:
            continue
        gs_total = sum(gs[c] or 0 for c in CATEGORIES)
        gs_comparison[eid] = {"gs_total": gs_total, "instructor_devs": {}}
        for uid in instructor_ids:
            if eid in matrix and uid in matrix[eid]:
                gs_comparison[eid]["instructor_devs"][uid] = round(
                    score_total(matrix[eid][uid]) - gs_total, 1
                )

    return {
        "instructor_stats": instructor_stats,
        "icc_per_category": icc_per_category,
        "overall_icc": overall_icc,
        "overall_mean": overall_mean,
        "heatmap": heatmap,
        "essay_ids": essay_ids,
        "essays": {e["id"]: e for e in essays},
        "gs_comparison": gs_comparison,
        "category_labels": CATEGORY_LABELS,
        "group_cat_means": group_cat_means,
    }


# ── Auth helpers ──────────────────────────────────────────────────────────────

def require_login(role=None):
    def decorator(f):
        from functools import wraps
        @wraps(f)
        def wrapped(*args, **kwargs):
            if "user_id" not in session:
                return redirect(url_for("login"))
            if role and session.get("user_role") != role:
                flash("You do not have permission to view that page.")
                return redirect(url_for("login"))
            return f(*args, **kwargs)
        return wrapped
    return decorator


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        user = get_user_by_email(email)
        if user and verify_password(user, password):
            session["user_id"] = user["id"]
            session["user_name"] = user["name"]
            session["user_role"] = user["role"]
            return redirect(
                url_for("admin_dashboard" if user["role"] == "admin" else "instructor_dashboard")
            )
        flash("Email or password not recognised. Please try again.")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ── Secure file serving ───────────────────────────────────────────────────────

@app.route("/essay-image/<path:filename>")
def essay_image(filename):
    if "user_id" not in session:
        abort(403)
    return send_from_directory(UPLOAD_DIR, secure_filename(filename))


@app.route("/essay-reading/<path:filename>")
def essay_reading(filename):
    if "user_id" not in session:
        abort(403)
    return send_from_directory(READING_DIR, secure_filename(filename))


@app.route("/training-file/<path:filename>")
def training_file(filename):
    if "user_id" not in session:
        abort(403)
    return send_from_directory(TRAINING_DIR, secure_filename(filename))


# ── Instructor routes ─────────────────────────────────────────────────────────

@app.route("/instructor")
@require_login()
def instructor_dashboard():
    if session.get("user_role") == "admin":
        return redirect(url_for("admin_dashboard"))

    uid = session["user_id"]
    phase = get_phase()

    generate_assignments_for_rater(uid)

    essays = get_all_essays()
    scored_ids = {s["essay_id"] for s in get_scores_for_user(uid)}
    assignments = get_rater_assignments(uid)
    total_assigned = len(assignments)
    done = sum(1 for a in assignments if a["essay_id"] in scored_ids)
    all_done = (total_assigned > 0 and done >= total_assigned)

    next_essay_id = get_next_essay_for_rater(uid) if phase == "data_collection" else None

    return render_template(
        "instructor/dashboard.html",
        phase=phase,
        next_essay_id=next_essay_id,
        total_assigned=total_assigned,
        done=done,
        all_done=all_done,
    )


@app.route("/instructor/essay/<int:essay_id>", methods=["GET", "POST"])
@require_login()
def mark_essay(essay_id):
    if session.get("user_role") == "admin":
        return redirect(url_for("admin_dashboard"))

    phase = get_phase()
    if phase != "data_collection":
        flash("Scoring is not currently open.")
        return redirect(url_for("instructor_dashboard"))

    uid = session["user_id"]
    essay = get_essay(essay_id)
    if not essay:
        flash("Sample not found.")
        return redirect(url_for("instructor_dashboard"))

    assignments = get_rater_assignments(uid)
    assigned_ids = [a["essay_id"] for a in assignments]
    if essay_id not in assigned_ids:
        flash("You are not assigned to this sample.")
        return redirect(url_for("instructor_dashboard"))

    existing = get_score_for_user_essay(uid, essay_id)

    if existing:
        flash("You have already submitted a score for this sample. Scores cannot be edited.")
        return redirect(url_for("instructor_dashboard"))

    next_id = get_next_essay_for_rater(uid)
    if next_id != essay_id:
        flash("Please complete the samples in the order they are presented.")
        return redirect(url_for("instructor_dashboard"))

    images = get_essay_images(essay_id)
    pos, total = get_rater_position_info(uid, essay_id)
    std = get_standardization(essay["standardization_id"]) if essay["standardization_id"] else None

    if request.method == "POST":
        errors, vals = [], {}
        for cat in CATEGORIES:
            scale = CATEGORY_SCALES[cat]
            try:
                v = int(request.form[cat])
                if not (scale["min"] <= v <= scale["max"]):
                    errors.append(f"{CATEGORY_LABELS[cat]}: value must be {scale['min']}–{scale['max']}.")
                vals[cat] = v
            except (KeyError, ValueError):
                errors.append(f"{CATEGORY_LABELS[cat]}: please select a score.")

        certainty = None
        try:
            certainty = int(request.form["certainty"])
            if not (1 <= certainty <= 5):
                errors.append("Certainty: please select a value between 1 and 5.")
        except (KeyError, ValueError):
            errors.append("Certainty: please indicate how certain you are of your overall score.")

        if errors:
            for e in errors:
                flash(e)
            return render_template(
                "instructor/essay.html",
                essay=essay, images=images,
                std=std,
                categories=CATEGORIES, category_labels=CATEGORY_LABELS,
                category_scales=CATEGORY_SCALES,
                certainty_labels=CERTAINTY_LABELS,
                existing=dict(request.form),
                pos=pos, total=total,
                max_total=MAX_TOTAL,
            )

        comments = request.form.get("comments", "").strip()
        conn = get_db()
        conn.execute("""
            INSERT INTO scores
                (user_id, essay_id, task_req, argument, support, language,
                 readability, formatting, certainty, comments)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (uid, essay_id,
              vals["task_req"], vals["argument"], vals["support"],
              vals["language"], vals["readability"], vals["formatting"],
              certainty, comments))
        conn.commit()
        conn.close()

        try:
            recompute_instructor_metrics()
        except Exception:
            pass

        flash(f"Sample {pos} of {total} submitted. Well done!")
        return redirect(url_for("instructor_dashboard"))

    return render_template(
        "instructor/essay.html",
        essay=essay, images=images,
        std=std,
        categories=CATEGORIES, category_labels=CATEGORY_LABELS,
        category_scales=CATEGORY_SCALES,
        certainty_labels=CERTAINTY_LABELS,
        existing=None,
        pos=pos, total=total,
        max_total=MAX_TOTAL,
    )


@app.route("/instructor/results")
@require_login()
def instructor_results():
    if session.get("user_role") == "admin":
        return redirect(url_for("admin_dashboard"))

    phase = get_phase()
    uid = session["user_id"]

    assignments = get_rater_assignments(uid)
    scored = get_scores_for_user(uid)
    scored_map = {s["essay_id"]: s for s in scored}

    if phase == "data_collection" and len(scored) < len(assignments):
        flash("Results are available once you have scored all samples, or when the admin opens the reporting phase.")
        return redirect(url_for("instructor_dashboard"))

    all_scores = get_all_scores()
    group_by_essay = {}
    for s in all_scores:
        group_by_essay.setdefault(s["essay_id"], []).append(s)

    results, my_totals = [], []
    for a in assignments:
        eid = a["essay_id"]
        my = scored_map.get(eid)
        if not my:
            continue
        my_total = score_total(my)
        my_totals.append(my_total)
        group = group_by_essay.get(eid, [])
        group_t = [score_total(s) for s in group]
        group_mean = round(statistics.mean(group_t), 1) if group_t else my_total
        gs = get_gold_standard(eid)
        gs_total = sum(gs[c] or 0 for c in CATEGORIES) if gs else None

        per_cat = []
        for cat in CATEGORIES:
            group_cat_vals = [s[cat] for s in group if s[cat] is not None]
            group_cat_mean = round(statistics.mean(group_cat_vals), 1) if group_cat_vals else None
            per_cat.append({
                "label": CATEGORY_LABELS[cat],
                "my": my[cat],
                "group_mean": group_cat_mean,
                "gs": gs[cat] if gs else None,
                "leniency": round((my[cat] or 0) - group_cat_mean, 2) if group_cat_mean is not None and my[cat] is not None else None,
            })

        results.append({
            "essay_id": eid,
            "my_total": my_total,
            "group_mean": group_mean,
            "gs_total": gs_total,
            "deviation": round(my_total - group_mean, 1),
            "per_cat": per_cat,
            "certainty": my["certainty"],
        })

    overall_my_mean = round(statistics.mean(my_totals), 2) if my_totals else 0
    instructor_means = {}
    for s in all_scores:
        instructor_means.setdefault(s["user_id"], []).append(score_total(s))
    completed = [statistics.mean(v) for v in instructor_means.values() if len(v) >= len(assignments)]
    group_overall_mean = round(statistics.mean(completed), 2) if completed else overall_my_mean
    leniency = round(overall_my_mean - group_overall_mean, 2)
    tendency = "Lenient" if leniency > 1.0 else ("Harsh" if leniency < -1.0 else "Calibrated")

    group_cat_means = {}
    for cat in CATEGORIES:
        vals = [s[cat] for s in all_scores if s[cat] is not None]
        group_cat_means[cat] = round(statistics.mean(vals), 2) if vals else None

    my_cat_means = {}
    for cat in CATEGORIES:
        vals = [scored_map[eid][cat] for eid in scored_map if scored_map[eid][cat] is not None]
        my_cat_means[cat] = round(statistics.mean(vals), 2) if vals else None

    cat_leniency = {}
    for cat in CATEGORIES:
        if my_cat_means.get(cat) is not None and group_cat_means.get(cat) is not None:
            cat_leniency[cat] = round(my_cat_means[cat] - group_cat_means[cat], 2)
        else:
            cat_leniency[cat] = None

    return render_template(
        "instructor/results.html",
        results=results,
        overall_my_mean=overall_my_mean,
        group_overall_mean=group_overall_mean,
        leniency=leniency,
        tendency=tendency,
        categories=CATEGORIES,
        category_labels=CATEGORY_LABELS,
        cat_leniency=cat_leniency,
        my_cat_means=my_cat_means,
        group_cat_means=group_cat_means,
        certainty_labels=CERTAINTY_LABELS,
    )


# ── Admin dashboard ───────────────────────────────────────────────────────────

@app.route("/admin")
@require_login(role="admin")
def admin_dashboard():
    essays = get_all_essays()
    instructors = get_all_instructors()
    submission_counts = {i["id"]: get_submission_count(i["id"]) for i in instructors}
    phase = get_phase()
    return render_template(
        "admin/dashboard.html",
        essays=essays, instructors=instructors,
        submission_counts=submission_counts,
        total_essays=len(essays),
        phase=phase, phases=PHASES,
    )


@app.route("/admin/phase", methods=["POST"])
@require_login(role="admin")
def admin_set_phase():
    phase = request.form.get("phase")
    if phase in PHASES:
        set_phase(phase)
        flash(f"Mode switched to: {PHASES[phase]}")
    return redirect(url_for("admin_dashboard"))


# ── Admin: essays ─────────────────────────────────────────────────────────────

@app.route("/admin/essays", methods=["GET", "POST"])
@require_login(role="admin")
def admin_essays():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        essay_number = request.form.get("essay_number", "").strip()
        notes = request.form.get("notes", "").strip()
        task_question = request.form.get("task_question", "").strip()
        author_tag = request.form.get("author_tag", "").strip() or generate_author_tag()
        source_instructor_bilkent_id = request.form.get("source_instructor_bilkent_id", "").strip() or None
        files = request.files.getlist("images")
        reading_file = request.files.get("reading_file")

        standardization_id = request.form.get("standardization_id", "").strip() or None
        if standardization_id:
            try:
                standardization_id = int(standardization_id)
            except ValueError:
                standardization_id = None

        if not title or not essay_number:
            flash("Essay number and title are required.")
        elif not files or all(f.filename == "" for f in files):
            flash("Please upload at least one image of the essay.")
        else:
            conn = get_db()
            system_id = generate_system_id()

            # Handle optional reading file
            reading_filename = None
            if reading_file and reading_file.filename and allowed_reading(reading_file.filename):
                ext = reading_file.filename.rsplit(".", 1)[1].lower()
                reading_filename = f"reading_{uuid.uuid4().hex[:10]}.{ext}"
                reading_file.save(os.path.join(READING_DIR, reading_filename))

            cur = conn.execute(
                """INSERT INTO essays
                   (essay_number, title, notes, task_question,
                    author_tag, source_instructor_bilkent_id, system_id, reading_filename,
                    standardization_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (int(essay_number), title, notes or None,
                 task_question or None, author_tag,
                 source_instructor_bilkent_id, system_id, reading_filename,
                 standardization_id),
            )
            essay_id = cur.lastrowid
            page = 1
            for f in files:
                if f and f.filename and allowed_file(f.filename):
                    ext = f.filename.rsplit(".", 1)[1].lower()
                    safe_name = f"essay{essay_id}_p{page}_{uuid.uuid4().hex[:8]}.{ext}"
                    f.save(os.path.join(UPLOAD_DIR, safe_name))
                    conn.execute(
                        "INSERT INTO essay_images (essay_id, filename, page_number) VALUES (?, ?, ?)",
                        (essay_id, safe_name, page),
                    )
                    page += 1
            conn.commit()
            conn.close()
            assign_new_essay_to_all_raters(essay_id)
            flash(f"Essay {essay_number} added (ID: {system_id}) with {page-1} file(s).")
        return redirect(url_for("admin_essays"))

    essays = get_all_essays()
    essay_images = {e["id"]: get_essay_images(e["id"]) for e in essays}
    instructors = get_all_instructors()
    standardizations = get_all_standardizations()
    return render_template(
        "admin/essays.html",
        essays=essays, essay_images=essay_images,
        instructors=instructors,
        standardizations=standardizations,
        now_year=datetime.now().year,
    )


@app.route("/admin/essays/<int:essay_id>/upload-reading", methods=["POST"])
@require_login(role="admin")
def upload_essay_reading(essay_id):
    essay = get_essay(essay_id)
    if not essay:
        flash("Essay not found.")
        return redirect(url_for("admin_essays"))
    reading_file = request.files.get("reading_file")
    if not reading_file or not reading_file.filename:
        flash("No file selected.")
        return redirect(url_for("admin_essays"))
    if not allowed_reading(reading_file.filename):
        flash("Allowed reading formats: PDF, DOC, DOCX, TXT, JPG, PNG.")
        return redirect(url_for("admin_essays"))

    # Remove old reading file if present
    if essay["reading_filename"]:
        try:
            os.remove(os.path.join(READING_DIR, essay["reading_filename"]))
        except FileNotFoundError:
            pass

    ext = reading_file.filename.rsplit(".", 1)[1].lower()
    reading_filename = f"reading_{uuid.uuid4().hex[:10]}.{ext}"
    reading_file.save(os.path.join(READING_DIR, reading_filename))

    conn = get_db()
    conn.execute("UPDATE essays SET reading_filename = ? WHERE id = ?", (reading_filename, essay_id))
    conn.commit()
    conn.close()
    flash(f"Assigned reading uploaded for Essay {essay['essay_number']}.")
    return redirect(url_for("admin_essays"))


@app.route("/admin/essays/<int:essay_id>/delete-reading", methods=["POST"])
@require_login(role="admin")
def delete_essay_reading(essay_id):
    essay = get_essay(essay_id)
    if not essay:
        flash("Essay not found.")
        return redirect(url_for("admin_essays"))
    if essay["reading_filename"]:
        try:
            os.remove(os.path.join(READING_DIR, essay["reading_filename"]))
        except FileNotFoundError:
            pass
        conn = get_db()
        conn.execute("UPDATE essays SET reading_filename = NULL WHERE id = ?", (essay_id,))
        conn.commit()
        conn.close()
        flash("Assigned reading removed.")
    return redirect(url_for("admin_essays"))


@app.route("/admin/essays/bulk-zip", methods=["POST"])
@require_login(role="admin")
def admin_bulk_zip():
    zfile = request.files.get("zipfile")
    if not zfile or not zfile.filename.lower().endswith(".zip"):
        flash("Please upload a .zip file.")
        return redirect(url_for("admin_essays"))

    try:
        zdata = zipfile.ZipFile(io.BytesIO(zfile.read()))
    except zipfile.BadZipFile:
        flash("Invalid ZIP file.")
        return redirect(url_for("admin_essays"))

    bulk_std_id = request.form.get("standardization_id", "").strip() or None
    if bulk_std_id:
        try:
            bulk_std_id = int(bulk_std_id)
        except ValueError:
            bulk_std_id = None

    conn = get_db()
    max_num = conn.execute("SELECT MAX(essay_number) as m FROM essays").fetchone()["m"] or 0
    conn.close()

    added, skipped = 0, []
    for zi in sorted(zdata.infolist(), key=lambda x: x.filename):
        name = zi.filename
        if zi.is_dir() or not name.lower().endswith(".pdf"):
            continue

        basename = os.path.splitext(os.path.basename(name))[0]
        parts = basename.split("_", 1)
        raw_author = parts[0]
        source_bilkent = parts[1].strip() if len(parts) > 1 else None
        author_tag = raw_author.replace("-", " ").strip()
        if not author_tag:
            author_tag = generate_author_tag()

        max_num += 1
        system_id = generate_system_id()
        title = f"Sample {max_num}"

        conn = get_db()
        try:
            cur = conn.execute(
                """INSERT INTO essays
                   (essay_number, title, author_tag, source_instructor_bilkent_id, system_id,
                    standardization_id)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (max_num, title, author_tag, source_bilkent, system_id, bulk_std_id),
            )
            essay_id = cur.lastrowid

            pdf_bytes = zdata.read(zi.filename)
            safe_name = f"essay{essay_id}_p1_{uuid.uuid4().hex[:8]}.pdf"
            with open(os.path.join(UPLOAD_DIR, safe_name), "wb") as fh:
                fh.write(pdf_bytes)
            conn.execute(
                "INSERT INTO essay_images (essay_id, filename, page_number) VALUES (?, ?, 1)",
                (essay_id, safe_name),
            )
            conn.commit()
            added += 1
        except Exception as ex:
            skipped.append(f"{basename}: {ex}")
            conn.rollback()
        finally:
            conn.close()

        if added:
            assign_new_essay_to_all_raters(essay_id)

    flash(f"Bulk upload complete: {added} sample(s) added." +
          (f" Skipped: {', '.join(skipped)}" if skipped else ""))
    return redirect(url_for("admin_essays"))


@app.route("/admin/essays/<int:essay_id>/add-image", methods=["POST"])
@require_login(role="admin")
def add_essay_image(essay_id):
    essay = get_essay(essay_id)
    if not essay:
        flash("Essay not found.")
        return redirect(url_for("admin_essays"))
    files = request.files.getlist("images")
    if not files or all(f.filename == "" for f in files):
        flash("No files selected.")
        return redirect(url_for("admin_essays"))
    conn = get_db()
    existing = get_essay_images(essay_id)
    page = (max(i["page_number"] for i in existing) + 1) if existing else 1
    added = 0
    for f in files:
        if f and f.filename and allowed_file(f.filename):
            ext = f.filename.rsplit(".", 1)[1].lower()
            safe_name = f"essay{essay_id}_p{page}_{uuid.uuid4().hex[:8]}.{ext}"
            f.save(os.path.join(UPLOAD_DIR, safe_name))
            conn.execute(
                "INSERT INTO essay_images (essay_id, filename, page_number) VALUES (?, ?, ?)",
                (essay_id, safe_name, page),
            )
            page += 1
            added += 1
    conn.commit()
    conn.close()
    flash(f"{added} file(s) added to Essay {essay['essay_number']}.")
    return redirect(url_for("admin_essays"))


@app.route("/admin/essays/<int:essay_id>/delete-image/<int:image_id>", methods=["POST"])
@require_login(role="admin")
def delete_essay_image(essay_id, image_id):
    conn = get_db()
    img = conn.execute(
        "SELECT * FROM essay_images WHERE id = ? AND essay_id = ?", (image_id, essay_id)
    ).fetchone()
    if img:
        try:
            os.remove(os.path.join(UPLOAD_DIR, img["filename"]))
        except FileNotFoundError:
            pass
        conn.execute("DELETE FROM essay_images WHERE id = ?", (image_id,))
        conn.commit()
        flash("File removed.")
    conn.close()
    return redirect(url_for("admin_essays"))


@app.route("/admin/essays/<int:essay_id>/edit", methods=["POST"])
@require_login(role="admin")
def edit_essay(essay_id):
    essay = get_essay(essay_id)
    if not essay:
        flash("Essay not found.")
        return redirect(url_for("admin_essays"))
    title = request.form.get("title", "").strip()
    essay_number = request.form.get("essay_number", "").strip()
    notes = request.form.get("notes", "").strip()
    task_question = request.form.get("task_question", "").strip()
    author_tag = request.form.get("author_tag", "").strip()
    source_instructor_bilkent_id = request.form.get("source_instructor_bilkent_id", "").strip() or None
    if not title or not essay_number:
        flash("Essay number and title are required.")
        return redirect(url_for("admin_essays"))
    conn = get_db()
    conn.execute(
        """UPDATE essays SET essay_number=?, title=?, notes=?, task_question=?,
           author_tag=?, source_instructor_bilkent_id=? WHERE id=?""",
        (int(essay_number), title, notes or None,
         task_question or None, author_tag or None,
         source_instructor_bilkent_id, essay_id),
    )
    conn.commit()
    conn.close()
    flash(f"Essay {essay_number} updated.")
    return redirect(url_for("admin_essays"))


@app.route("/admin/essays/<int:essay_id>/delete", methods=["POST"])
@require_login(role="admin")
def delete_essay(essay_id):
    conn = get_db()
    essay = get_essay(essay_id)
    if essay and essay["reading_filename"]:
        try:
            os.remove(os.path.join(READING_DIR, essay["reading_filename"]))
        except FileNotFoundError:
            pass
    images = conn.execute(
        "SELECT filename FROM essay_images WHERE essay_id = ?", (essay_id,)
    ).fetchall()
    for img in images:
        try:
            os.remove(os.path.join(UPLOAD_DIR, img["filename"]))
        except FileNotFoundError:
            pass
    conn.execute("DELETE FROM essay_images WHERE essay_id = ?", (essay_id,))
    conn.execute("DELETE FROM scores WHERE essay_id = ?", (essay_id,))
    conn.execute("DELETE FROM gold_standards WHERE essay_id = ?", (essay_id,))
    conn.execute("DELETE FROM rater_assignments WHERE essay_id = ?", (essay_id,))
    conn.execute("DELETE FROM essays WHERE id = ?", (essay_id,))
    conn.commit()
    conn.close()
    flash("Sample and all associated data deleted.")
    return redirect(url_for("admin_essays"))


# ── Admin: standardized scores ────────────────────────────────────────────────

@app.route("/admin/gold-standard", methods=["GET", "POST"])
@require_login(role="admin")
def admin_gold_standard():
    essays = get_all_essays()
    standardizations = get_all_standardizations()
    # Group essays by standardization_id for the sidebar
    essays_by_std = {}
    for e in essays:
        key = e["standardization_id"]
        essays_by_std.setdefault(key, []).append(e)
    if request.method == "POST":
        essay_id = int(request.form.get("essay_id"))
        vals, errors = {}, []
        for cat in CATEGORIES:
            scale = CATEGORY_SCALES[cat]
            try:
                v = int(request.form[cat])
                if not (scale["min"] <= v <= scale["max"]):
                    errors.append(f"{CATEGORY_LABELS[cat]} must be {scale['min']}–{scale['max']}.")
                vals[cat] = v
            except (KeyError, ValueError):
                errors.append(f"Missing value for {CATEGORY_LABELS[cat]}.")

        # Optional variance and std_dev
        score_variance = None
        score_std_dev = None
        try:
            v_str = request.form.get("score_variance", "").strip()
            if v_str:
                score_variance = float(v_str)
        except ValueError:
            errors.append("Variance must be a number.")
        try:
            sd_str = request.form.get("score_std_dev", "").strip()
            if sd_str:
                score_std_dev = float(sd_str)
        except ValueError:
            errors.append("Standard deviation must be a number.")

        if errors:
            for e in errors:
                flash(e)
        else:
            conn = get_db()
            conn.execute("""
                INSERT INTO gold_standards
                    (essay_id, task_req, argument, support, language, readability, formatting,
                     score_variance, score_std_dev, set_by)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(essay_id) DO UPDATE SET
                    task_req=excluded.task_req, argument=excluded.argument,
                    support=excluded.support, language=excluded.language,
                    readability=excluded.readability, formatting=excluded.formatting,
                    score_variance=excluded.score_variance,
                    score_std_dev=excluded.score_std_dev,
                    set_by=excluded.set_by, set_at=CURRENT_TIMESTAMP
            """, (essay_id,
                  vals["task_req"], vals["argument"], vals["support"],
                  vals["language"], vals["readability"], vals["formatting"],
                  score_variance, score_std_dev,
                  session["user_id"]))
            conn.commit()
            conn.close()
            flash("Standardized score saved.")
        return redirect(url_for("admin_gold_standard"))
    gold_standards = {e["id"]: get_gold_standard(e["id"]) for e in essays}
    return render_template(
        "admin/gold_standard.html",
        essays=essays, gold_standards=gold_standards,
        standardizations=standardizations,
        essays_by_std=essays_by_std,
        categories=CATEGORIES, category_labels=CATEGORY_LABELS,
        category_scales=CATEGORY_SCALES,
    )


# ── Admin: instructors ────────────────────────────────────────────────────────

@app.route("/admin/instructors", methods=["GET", "POST"])
@require_login(role="admin")
def admin_instructors():
    if request.method == "POST":
        action = request.form.get("action", "add")

        if action == "reset_password":
            uid = request.form.get("user_id")
            pw = request.form.get("new_password", "").strip()
            if not pw or len(pw) < 6:
                flash("Password must be at least 6 characters.")
            else:
                conn = get_db()
                conn.execute(
                    "UPDATE users SET password_hash = ? WHERE id = ? AND role = 'instructor'",
                    (generate_password_hash(pw), uid),
                )
                conn.commit()
                conn.close()
                flash("Password updated.")
            return redirect(url_for("admin_instructors"))

        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "").strip()
        bilkent_id = request.form.get("bilkent_id", "").strip() or None
        if not name or not email or not password:
            flash("Name, email, and password are all required.")
        elif len(password) < 6:
            flash("Password must be at least 6 characters.")
        else:
            conn = get_db()
            if conn.execute("SELECT id FROM users WHERE email = ?", (email,)).fetchone():
                flash("An account with that email already exists.")
            else:
                cur = conn.execute(
                    "INSERT INTO users (name, email, password_hash, role, bilkent_id) VALUES (?, ?, ?, 'instructor', ?)",
                    (name, email, generate_password_hash(password), bilkent_id),
                )
                new_uid = cur.lastrowid
                conn.commit()
                conn.close()
                generate_assignments_for_rater(new_uid)
                flash(f"{name} added. They can log in with {email}.")
                return redirect(url_for("admin_instructors"))
            conn.close()
        return redirect(url_for("admin_instructors"))

    instructors = get_all_instructors()
    submission_counts = {i["id"]: get_submission_count(i["id"]) for i in instructors}
    total_essays = len(get_all_essays())
    return render_template(
        "admin/instructors.html",
        instructors=instructors,
        submission_counts=submission_counts,
        total_essays=total_essays,
    )


@app.route("/admin/instructors/csv-upload", methods=["POST"])
@require_login(role="admin")
def admin_instructors_csv():
    csvfile = request.files.get("csvfile")
    mode = request.form.get("mode", "add_only")
    if not csvfile or not csvfile.filename.lower().endswith(".csv"):
        flash("Please upload a .csv file.")
        return redirect(url_for("admin_instructors"))

    try:
        content = csvfile.read().decode("utf-8-sig")
        reader = csv.DictReader(content.splitlines())
    except Exception as ex:
        flash(f"Could not parse CSV: {ex}")
        return redirect(url_for("admin_instructors"))

    required_cols = {"bilkent_id", "name", "email"}
    if not reader.fieldnames or not required_cols.issubset(set(reader.fieldnames)):
        flash(f"CSV must have columns: {', '.join(required_cols)} (plus optional 'password').")
        return redirect(url_for("admin_instructors"))

    rows = list(reader)
    csv_bilkent_ids = set()
    added, skipped = 0, 0

    conn = get_db()
    for row in rows:
        bid = row.get("bilkent_id", "").strip()
        name = row.get("name", "").strip()
        email = row.get("email", "").strip().lower()
        password = row.get("password", "").strip()
        if not bid or not name or not email:
            skipped += 1
            continue
        csv_bilkent_ids.add(bid)
        if not password:
            import secrets, string
            password = "".join(secrets.choice(string.ascii_letters + string.digits) for _ in range(10))

        existing = conn.execute(
            "SELECT id FROM users WHERE bilkent_id = ? OR email = ?", (bid, email)
        ).fetchone()
        if existing:
            skipped += 1
            continue

        cur = conn.execute(
            "INSERT INTO users (name, email, password_hash, role, bilkent_id) VALUES (?, ?, ?, 'instructor', ?)",
            (name, email, generate_password_hash(password), bid),
        )
        new_uid = cur.lastrowid
        added += 1
        conn.commit()
        generate_assignments_for_rater(new_uid)

    if mode == "remove_leavers" and csv_bilkent_ids:
        all_instructors = conn.execute(
            "SELECT id, bilkent_id FROM users WHERE role = 'instructor' AND bilkent_id IS NOT NULL"
        ).fetchall()
        removed = 0
        for instr in all_instructors:
            if instr["bilkent_id"] not in csv_bilkent_ids:
                conn.execute("DELETE FROM scores WHERE user_id = ?", (instr["id"],))
                conn.execute("DELETE FROM rater_assignments WHERE user_id = ?", (instr["id"],))
                conn.execute("DELETE FROM users WHERE id = ?", (instr["id"],))
                removed += 1
        conn.commit()
        flash(f"CSV import: {added} added, {skipped} skipped, {removed} removed (not in CSV).")
    else:
        conn.commit()
        flash(f"CSV import: {added} added, {skipped} skipped (already exist).")

    conn.close()
    return redirect(url_for("admin_instructors"))


@app.route("/admin/instructors/<int:user_id>/delete", methods=["POST"])
@require_login(role="admin")
def delete_instructor(user_id):
    conn = get_db()
    conn.execute("DELETE FROM scores WHERE user_id = ?", (user_id,))
    conn.execute("DELETE FROM rater_assignments WHERE user_id = ?", (user_id,))
    conn.execute("DELETE FROM users WHERE id = ? AND role = 'instructor'", (user_id,))
    conn.commit()
    conn.close()
    flash("Instructor removed.")
    return redirect(url_for("admin_instructors"))


@app.route("/admin/instructors/recalculate", methods=["POST"])
@require_login(role="admin")
def admin_recalculate_metrics():
    recompute_instructor_metrics()
    flash("Instructor metrics recalculated.")
    return redirect(url_for("admin_instructor_analysis"))


# ── Admin: analysis dashboards ────────────────────────────────────────────────

@app.route("/admin/sample-analysis")
@require_login(role="admin")
def admin_sample_analysis():
    stats = compute_reliability_stats()
    return render_template(
        "admin/sample_analysis.html",
        stats=stats,
        categories=CATEGORIES,
        category_labels=CATEGORY_LABELS,
        max_total=MAX_TOTAL,
    )


@app.route("/admin/instructor-analysis")
@require_login(role="admin")
def admin_instructor_analysis():
    stats = compute_reliability_stats()
    instructors = get_all_instructors()
    return render_template(
        "admin/instructor_analysis.html",
        stats=stats,
        instructors=instructors,
        categories=CATEGORIES,
        category_labels=CATEGORY_LABELS,
        certainty_labels=CERTAINTY_LABELS,
    )


@app.route("/admin/mfrm")
@require_login(role="admin")
def admin_mfrm():
    all_scores = get_all_scores()
    essays     = get_all_essays()
    instructors = get_all_instructors()

    # Build score matrix {essay_id: {rater_id: total_score}}
    score_matrix = {}
    for s in all_scores:
        total = score_total(s)
        score_matrix.setdefault(s["essay_id"], {})[s["user_id"]] = total

    mfrm_results = compute_mfrm(score_matrix, max_score=MAX_TOTAL) if len(score_matrix) >= 2 else None

    instr_map = {i["id"]: i for i in instructors}
    essay_map = {e["id"]: e for e in essays}

    return render_template(
        "admin/mfrm_analysis.html",
        mfrm=mfrm_results,
        instr_map=instr_map,
        essay_map=essay_map,
        max_total=MAX_TOTAL,
    )


@app.route("/admin/reliability")
@require_login(role="admin")
def admin_reliability():
    return redirect(url_for("admin_sample_analysis"))


@app.route("/admin/reliability/data")
@require_login(role="admin")
def reliability_data_json():
    stats = compute_reliability_stats()
    if not stats:
        return jsonify({})
    return jsonify({
        "instructor_names": [v["name"] for v in stats["instructor_stats"].values()],
        "leniency_values": [v["leniency"] for v in stats["instructor_stats"].values()],
        "icc_categories": {CATEGORY_LABELS[k]: v for k, v in stats["icc_per_category"].items()},
        "overall_icc": stats["overall_icc"],
    })


# ── Admin: Excel export ───────────────────────────────────────────────────────

@app.route("/admin/export/excel")
@require_login(role="admin")
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash("openpyxl not installed. Run: pip install openpyxl")
        return redirect(url_for("admin_sample_analysis"))

    all_scores = get_all_scores()
    essays = get_all_essays()
    instructors = get_all_instructors()
    matrix = {}
    for s in all_scores:
        matrix.setdefault(s["essay_id"], {})[s["user_id"]] = s
    essays_dict = {e["id"]: e for e in essays}

    wb = Workbook()
    hfil = PatternFill("solid", fgColor="1C1C1C")
    hfnt = Font(color="FFFFFF", bold=True)
    hctr = Alignment(horizontal="center")

    def style_header(ws, headers):
        ws.append(headers)
        for i in range(1, len(headers) + 1):
            c = ws.cell(row=1, column=i)
            c.fill = hfil; c.font = hfnt; c.alignment = hctr

    def autowidth(ws, cap=40):
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max((len(str(c.value)) if c.value else 0) for c in col) + 4, cap,
            )

    # Sheet 1: All Scores
    ws1 = wb.active; ws1.title = "All Scores"
    style_header(ws1, [
        "System ID", "Essay #", "Author Tag", "Instructor", "Bilkent ID",
        "Task Requirements", "Argument", "Support", "Use of Language",
        "Readability", "Formatting", "Total (/22)", "Certainty", "Comments", "Submitted At",
    ])
    for s in all_scores:
        eq = essays_dict.get(s["essay_id"])
        ws1.append([
            eq["system_id"] if eq else "",
            s["essay_number"],
            s["author_tag"] or "",
            s["instructor_name"],
            s["instructor_bilkent_id"] or "",
            s["task_req"], s["argument"], s["support"], s["language"],
            s["readability"], s["formatting"], score_total(s),
            s["certainty"], s["comments"] or "", str(s["submitted_at"]),
        ])
    autowidth(ws1)

    # Sheet 2: Summary by Essay
    ws2 = wb.create_sheet("Summary by Essay")
    h2 = ["System ID", "Essay #", "Author Tag", "Standardized Score (/22)", "Variance", "Std Dev"]
    for instr in instructors:
        h2.append(instr["name"])
    h2 += ["Group Mean", "Group Std Dev"]
    style_header(ws2, h2)
    for e in essays:
        eid = e["id"]
        gs = get_gold_standard(eid)
        gs_t = sum(gs[c] or 0 for c in CATEGORIES) if gs else ""
        gs_var = gs["score_variance"] if gs and gs["score_variance"] is not None else ""
        gs_sd = gs["score_std_dev"] if gs and gs["score_std_dev"] is not None else ""
        row = [e["system_id"] or "", e["essay_number"], e["author_tag"] or "", gs_t, gs_var, gs_sd]
        tots = []
        for instr in instructors:
            uid = instr["id"]
            if eid in matrix and uid in matrix[eid]:
                t = score_total(matrix[eid][uid]); row.append(t); tots.append(t)
            else:
                row.append("")
        row.append(round(statistics.mean(tots), 2) if tots else "")
        row.append(round(statistics.stdev(tots), 2) if len(tots) > 1 else (0.0 if tots else ""))
        ws2.append(row)
    autowidth(ws2, cap=35)

    # Sheet 3: Instructor Stats
    ws3 = wb.create_sheet("Instructor Stats")
    style_header(ws3, [
        "Instructor", "Bilkent ID", "Email", "Essays Marked",
        "Mean Total (/22)", "Leniency Index", "Tendency",
        "Avg Certainty", "Avg Distance from Standardized Score",
    ])
    rstats = compute_reliability_stats()
    if rstats:
        for uid, s in rstats["instructor_stats"].items():
            instr = next((i for i in instructors if i["id"] == uid), None)
            ws3.append([
                s["name"], s.get("bilkent_id", ""), s.get("email", ""),
                s["essays_marked"], s["mean_total"], s["leniency"], s["tendency"],
                s.get("avg_certainty", ""),
                instr["avg_distance"] if instr else "",
            ])
        ws3.append([])
        ws3.append(["ICC Summary"])
        ws3.append(["Category", "ICC", "Interpretation"])
        for cat in CATEGORIES:
            icc = rstats["icc_per_category"].get(cat)
            interp = ("Insufficient data" if icc is None else
                      "Good" if icc >= 0.75 else
                      "Acceptable" if icc >= 0.60 else "Needs calibration")
            ws3.append([CATEGORY_LABELS[cat], icc if icc is not None else "N/A", interp])
        ov = rstats.get("overall_icc")
        ovi = ("Insufficient data" if ov is None else
               "Good" if ov >= 0.75 else
               "Acceptable" if ov >= 0.60 else "Needs calibration")
        ws3.append(["Overall (total scores)", ov if ov is not None else "N/A", ovi])
    autowidth(ws3, cap=35)

    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname = f"calibration_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        out,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname,
    )


# ── Admin: standardizations ───────────────────────────────────────────────────

@app.route("/admin/standardizations", methods=["GET", "POST"])
@require_login(role="admin")
def admin_standardizations():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        year = request.form.get("year", "").strip()
        term = request.form.get("term", "").strip()
        reading_task = request.form.get("reading_task", "").strip()
        reading_file = request.files.get("reading_file")

        if not name or not year:
            flash("Name and year are required.")
            return redirect(url_for("admin_standardizations"))

        try:
            year = int(year)
        except ValueError:
            flash("Year must be a number.")
            return redirect(url_for("admin_standardizations"))

        reading_filename = None
        if reading_file and reading_file.filename and allowed_reading(reading_file.filename):
            ext = reading_file.filename.rsplit(".", 1)[1].lower()
            reading_filename = f"reading_{uuid.uuid4().hex[:10]}.{ext}"
            reading_file.save(os.path.join(READING_DIR, reading_filename))

        std_id = create_standardization(
            name, year, term or None,
            reading_task or None, reading_filename,
        )
        flash(f"Standardization '{name}' created.")
        return redirect(url_for("admin_standardization_detail", std_id=std_id))

    standardizations = get_all_standardizations()
    return render_template(
        "admin/standardizations.html",
        standardizations=standardizations,
        current_year=datetime.now().year,
    )


@app.route("/admin/standardizations/<int:std_id>")
@require_login(role="admin")
def admin_standardization_detail(std_id):
    std = get_standardization(std_id)
    if not std:
        flash("Standardization not found.")
        return redirect(url_for("admin_standardizations"))
    essays = get_essays_by_standardization(std_id)
    essay_ids = [e["id"] for e in essays]

    # Per-essay: gold standard, score count, group mean
    essay_data = []
    all_scores = get_all_scores()
    score_map = {}
    for s in all_scores:
        score_map.setdefault(s["essay_id"], []).append(s)

    for e in essays:
        eid = e["id"]
        gs = get_gold_standard(eid)
        gs_total = sum(gs[c] or 0 for c in CATEGORIES) if gs else None
        group_scores = score_map.get(eid, [])
        group_totals = [score_total(s) for s in group_scores]
        essay_data.append({
            "essay": e,
            "gs_total": gs_total,
            "score_count": len(group_scores),
            "group_mean": round(statistics.mean(group_totals), 1) if group_totals else None,
        })

    # Instructor stats scoped to this standardization
    cats = CATEGORIES
    by_instr = {}
    for s in all_scores:
        if s["essay_id"] in essay_ids:
            uid = s["user_id"]
            by_instr.setdefault(uid, {
                "name": s["instructor_name"],
                "bilkent_id": s["instructor_bilkent_id"] or "",
                "scores": [],
            })["scores"].append(s)

    instr_stats = []
    all_totals_here = [score_total(s) for s in all_scores if s["essay_id"] in set(essay_ids)]
    group_mean = statistics.mean(all_totals_here) if all_totals_here else 0
    for uid, data in by_instr.items():
        totals = [score_total(s) for s in data["scores"]]
        mean_t = statistics.mean(totals)
        leniency = round(mean_t - group_mean, 2)
        dists = []
        for s in data["scores"]:
            gs = get_gold_standard(s["essay_id"])
            if gs:
                gs_t = sum(gs[c] or 0 for c in cats)
                dists.append(abs(score_total(s) - gs_t))
        instr_stats.append({
            "id": uid,
            "name": data["name"],
            "bilkent_id": data["bilkent_id"],
            "essays_scored": len(data["scores"]),
            "mean_total": round(mean_t, 2),
            "leniency": leniency,
            "tendency": "Lenient" if leniency > 1.0 else ("Harsh" if leniency < -1.0 else "Calibrated"),
            "avg_distance": round(statistics.mean(dists), 2) if dists else None,
        })
    instr_stats.sort(key=lambda x: x["name"])

    # Essays not in this standardization (available to add)
    all_essays = get_all_essays()
    essay_ids_set = set(essay_ids)
    available_essays = [e for e in all_essays if e["id"] not in essay_ids_set]

    return render_template(
        "admin/standardization_detail.html",
        std=std,
        essay_data=essay_data,
        instr_stats=instr_stats,
        group_mean=round(group_mean, 2),
        categories=CATEGORIES,
        category_labels=CATEGORY_LABELS,
        available_essays=available_essays,
    )


@app.route("/admin/standardizations/<int:std_id>/remove-essay/<int:essay_id>", methods=["POST"])
@require_login(role="admin")
def admin_remove_essay_from_standardization(std_id, essay_id):
    conn = get_db()
    conn.execute(
        "UPDATE essays SET standardization_id = NULL WHERE id = ? AND standardization_id = ?",
        (essay_id, std_id),
    )
    conn.commit()
    conn.close()
    flash("Sample removed from this standardization.")
    return redirect(url_for("admin_standardization_detail", std_id=std_id))


@app.route("/admin/essays/bulk-assign", methods=["POST"])
@require_login(role="admin")
def admin_bulk_assign_essays():
    essay_ids_raw = request.form.getlist("essay_ids")
    essay_ids = [int(x) for x in essay_ids_raw if x.isdigit()]
    if not essay_ids:
        flash("No samples selected.")
        return redirect(url_for("admin_essays"))

    action = request.form.get("assign_action", "existing")

    if action == "new":
        name = request.form.get("new_name", "").strip()
        year = request.form.get("new_year", "").strip()
        term = request.form.get("new_term", "").strip()
        reading_task = request.form.get("new_reading_task", "").strip()
        if not name or not year:
            flash("Name and year are required for a new standardization.")
            return redirect(url_for("admin_essays"))
        try:
            year = int(year)
        except ValueError:
            flash("Year must be a number.")
            return redirect(url_for("admin_essays"))
        std_id = create_standardization(name, year, term or None, reading_task or None, None)
    else:
        std_id_raw = request.form.get("standardization_id", "").strip()
        if not std_id_raw:
            flash("Please select a standardization.")
            return redirect(url_for("admin_essays"))
        try:
            std_id = int(std_id_raw)
        except ValueError:
            flash("Invalid standardization.")
            return redirect(url_for("admin_essays"))

    conn = get_db()
    for eid in essay_ids:
        conn.execute("UPDATE essays SET standardization_id = ? WHERE id = ?", (std_id, eid))
    conn.commit()
    conn.close()
    flash(f"{len(essay_ids)} sample(s) assigned to standardization.")
    return redirect(url_for("admin_standardization_detail", std_id=std_id))


@app.route("/admin/standardizations/<int:std_id>/add-essays", methods=["POST"])
@require_login(role="admin")
def admin_add_essays_to_standardization(std_id):
    essay_ids_raw = request.form.getlist("essay_ids")
    essay_ids = [int(x) for x in essay_ids_raw if x.isdigit()]
    if not essay_ids:
        flash("No samples selected.")
        return redirect(url_for("admin_standardization_detail", std_id=std_id))
    conn = get_db()
    for eid in essay_ids:
        conn.execute("UPDATE essays SET standardization_id = ? WHERE id = ?", (std_id, eid))
    conn.commit()
    conn.close()
    flash(f"{len(essay_ids)} sample(s) added.")
    return redirect(url_for("admin_standardization_detail", std_id=std_id))


@app.route("/admin/standardizations/<int:std_id>/archive", methods=["POST"])
@require_login(role="admin")
def admin_archive_standardization(std_id):
    archive_standardization(std_id)
    flash("Standardization archived.")
    return redirect(url_for("admin_standardizations"))


@app.route("/admin/standardizations/<int:std_id>/reopen", methods=["POST"])
@require_login(role="admin")
def admin_reopen_standardization(std_id):
    reopen_standardization(std_id)
    flash("Standardization reopened.")
    return redirect(url_for("admin_standardizations"))


@app.route("/admin/standardizations/<int:std_id>/delete", methods=["POST"])
@require_login(role="admin")
def admin_delete_standardization(std_id):
    conn = get_db()
    conn.execute("DELETE FROM scores WHERE essay_id IN (SELECT id FROM essays WHERE standardization_id = ?)", (std_id,))
    conn.execute("DELETE FROM essay_images WHERE essay_id IN (SELECT id FROM essays WHERE standardization_id = ?)", (std_id,))
    conn.execute("DELETE FROM rater_assignments WHERE essay_id IN (SELECT id FROM essays WHERE standardization_id = ?)", (std_id,))
    conn.execute("DELETE FROM essays WHERE standardization_id = ?", (std_id,))
    conn.execute("DELETE FROM standardizations WHERE id = ?", (std_id,))
    conn.commit()
    conn.close()
    flash("Standardization deleted.")
    return redirect(url_for("admin_standardizations"))


@app.route("/admin/cross-standardization")
@require_login(role="admin")
def admin_cross_standardization():
    data = get_cross_standardization_stats()
    return render_template(
        "admin/cross_standardization.html",
        data=data,
    )


# ── Training ──────────────────────────────────────────────────────────────────

@app.route("/training")
@require_login()
def training():
    materials = get_training_materials()
    notes = get_training_notes()
    return render_template("training.html", materials=materials, notes=notes)


@app.route("/training/upload", methods=["POST"])
@require_login(role="admin")
def training_upload():
    f = request.files.get("file")
    description = request.form.get("description", "").strip()
    if not f or not f.filename:
        flash("No file selected.")
        return redirect(url_for("training"))
    if not allowed_training(f.filename):
        flash("File type not allowed.")
        return redirect(url_for("training"))
    import os as _os
    _os.makedirs(TRAINING_DIR, exist_ok=True)
    ext = f.filename.rsplit(".", 1)[1].lower()
    stored_name = f"training_{uuid.uuid4().hex[:12]}.{ext}"
    f.save(_os.path.join(TRAINING_DIR, stored_name))
    ftype = training_file_type(f.filename)
    add_training_material(stored_name, f.filename, ftype, description or None)
    flash(f"'{f.filename}' uploaded.")
    return redirect(url_for("training"))


@app.route("/training/materials/<int:material_id>/delete", methods=["POST"])
@require_login(role="admin")
def training_delete_material(material_id):
    filename = delete_training_material(material_id)
    if filename:
        import os as _os
        path = _os.path.join(TRAINING_DIR, filename)
        if _os.path.exists(path):
            _os.remove(path)
    flash("Material deleted.")
    return redirect(url_for("training"))


@app.route("/training/notes", methods=["POST"])
@require_login(role="admin")
def training_add_note():
    content = request.form.get("content", "").strip()
    color = request.form.get("color", "yellow")
    if not content:
        flash("Note cannot be empty.")
        return redirect(url_for("training"))
    add_training_note(content, color)
    return redirect(url_for("training"))

